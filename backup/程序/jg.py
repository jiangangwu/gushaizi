""" Wu Jiangang's functions

It's a collection of functions that Wu often use.
"""
def clearFileName(fileName):
    '''使文件名符合windows的要求'''
    fileName = fileName.replace('*','星号')\
               .replace('<','《')\
               .replace('>','》')\
               .replace('/','_')\
               .replace('|','_')\
               .replace('\n','')\
               .replace('\\','_')\
               .replace('"',"'")\
               .replace('?','？')\
               .replace(':','：')
    if len(fileName) > 221:
        fileName = fileName[:200]+"__" + fileName[-20:]
    return fileName

def date2Standard(date):
    '''使2012/2/8变成2012/02/02'''
    d = date.split('/')
    d2 = d[0]
    if len(d[1])==2:
        d2 = d2 + d[1]
    else:
        d2 = d2 + '0' + d[1]
    if len(d[2])==2:
        d2 = d2 + d[2]
    else:
        d2 = d2 + '0' + d[2]
    return d2

def time2second(x):
    '''使形如 2015-12-15 15：15：12 的时间变成1900-01-01 00：00：00 开始计算的秒数'''
    import time
    if '-' in x:
        a = time.mktime(time.strptime(x.replace('：',':').replace('-','-'), '%Y-%m-%d %H:%M:%S'))
    if '_' in x:
        a = time.mktime(time.strptime(x.replace('：',':'), '%Y_%m_%d %H:%M:%S'))
    if '/' in x:
        a = time.mktime(time.strptime(x.replace('：',':'), '%Y/%m/%d %H:%M:%S'))
    return a


def getTime():
    '''获取当时的时间用 2015-12-08 14：02：14 的形式表示出来。'''
    import time
    t = time.localtime()
    date = str(t.tm_year)
    if len(str(t.tm_mon))==2:
        date = date + '-' + str(t.tm_mon)
    else:
        date = date + '-' + '0' + str(t.tm_mon)
    if len(str(t.tm_mday))==2:
        date = date + '-' + str(t.tm_mday)
    else:
        date = date + '-' + '0' + str(t.tm_mday)
    if len(str(t.tm_hour))==2:
        date = date + ' ' + str(t.tm_hour)
    else:
        date = date + ' ' + '0' + str(t.tm_hour)
    if len(str(t.tm_min))==2:
        date = date + ':' + str(t.tm_min)
    else:
        date = date + ':' + '0' + str(t.tm_sec)
    if len(str(t.tm_sec))==2:
        date = date + ':' + str(t.tm_sec)
    else:
        date = date + ':' + '0' + str(t.tm_sec)
    return date

def getDate():
    '''获取当时的时间用 2015-12-08 的形式表示出来。'''
    import time
    t = time.localtime()
    date = str(t.tm_year)
    if len(str(t.tm_mon))==2:
        date = date + '-' + str(t.tm_mon)
    else:
        date = date + '-' + '0' + str(t.tm_mon)
    if len(str(t.tm_mday))==2:
        date = date + '-' + str(t.tm_mday)
    else:
        date = date + '-' + '0' + str(t.tm_mday)
    return date

def clrHtml(txt):
    '''清理html字符串中的<>及中间的部分，也清理空格和换行符。'''
    t = txt.replace('<','>').split('>')
    r = ''
    for i in range(0,len(t),2):
        r = r + t[i]
    r = r.replace(' ','').replace('\r','').replace('\t','').replace('\n','')
    return r

def date(days):
    """date(days)：取时间，参数可以取向前推或后推几天的时间"""
    import time
    t = time.localtime(time.time() + (days + 1) * 86400)
    d = str(t.tm_year) + '-' + str(t.tm_mon) + '-' + str(t.tm_mday-1)
    return d

def now():
    import time
    return time.ctime()[11:20]

def day():
    import time
    return time.localtime().tm_mday

def week():
    import time
    return time.localtime().tm_wday

def month():
    import time
    return time.localtime().tm_mon

def deleteFile(file):
    """dltFl(file):删除指定文件或目录"""
    import os
    if os.path.isfile(file):
          os.remove(file)
          print('File deleted!')
    elif os.path.isdir(file):
          os.rmdir(file)
          print('Dir deleted!')
    else:
          print('The file or dir doesn\'t exist in the dir.')

def copyFile(srcFile, dst):
    import shutil
    shutil.copy(srcFile, dst)

def moveFile(file, path):
    import os, shutil
    shutil.move(path, ne)

def dir2lst(dir):
    """把指定目录下的文件传入列表"""
    import os
    fl = []
    for root, dirs, files in os.walk(dir, True):
        for name in files:
            fl.append(os.path.join(root,name))
    return fl

def downloadWindReport(SocketTime,urlOpenTimeOut,\
                         urls,tag1, tag2, localStoreDir, fileNames):
    """
    dnldRpt(SocketTime,urlOpenTimeOut,urls,tag1, tag2, localStoreDir, fileNames):
    从万得下载公告的函数
    urls,fileNames是列表，可以在EXCEL里加工后导入，再输入到函数
    建议：SocketTime = 300，urlOpenTimeOut = 10
    例：
    LocalStoreDir = 'C:/JR/招股说明书/招股说明书PDF编号/'
    tag1 = "<a href="
    tag2 = " class='big'  title="
    """
    import os, socket, urllib.request
    socket.setdefaulttimeout(SocketTime)
    for i in range(0,len(urls)):
       try:
           urlopen = urllib.request.urlopen(urls[i],timeout = urlOpenTimeOut)
           html = urlopen.read().decode('utf-8')
           fileLink = html.split(tag1)[1].split(tag2)[0]
           urllib.request.urlretrieve(fileLink, localStoreDir + fileNames[i] )
       except Exception as e:
           print(e)
           if os.path.isfile(localStoreDir + fileNames[i]):
               os.remove(localStoreDir + fileNames[i])
               print('removed the unfinished file!')
           continue
       print(str(i) + ':' + fileNames[i])


        
def xl2lst(xlFile, clmnNo):
    """读入Excel文件的指定列到列表中"""
    import xlrd
    data = xlrd.open_workbook(xlFile)
    table = data.sheets()[0]
    lst = table.col_values(clmnNo-1)
    return lst

      
def lst2xl(records, fileName, row):
    '''将list写入excel2007   需要指定写入的行'''
    from openpyxl.workbook import Workbook     
    from openpyxl.writer.excel import ExcelWriter  
    from openpyxl.cell import get_column_letter 
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    ws = wb.worksheets[0]
    ws.title = '1'
    for record in records:  
        for x in range(1,len(record)+1):  
            col = get_column_letter(x)  
            ws.cell('%s%s'%(col, row)).value = '%s' % (record[x-1])        
        row += 1     
    ew.save(filename = fileName) 

def lst2xlWithHeaders(records, fileName, row, headers):
    """将list写入excel2007   需要为第一行指定表头
    格式：headers = ['id','Stkcd','Reptdt']"""
    from openpyxl.workbook import Workbook     
    from openpyxl.writer.excel import ExcelWriter  
    from openpyxl.cell import get_column_letter 
    wb = Workbook()
    ew = ExcelWriter(workbook = wb)
    ws = wb.worksheets[0]
    ws.title = '1'
    x = 1
    i = 1
    for x in range(1,len(headers)+1):  
        col = get_column_letter(x)  
        ws.cell('%s%s'%(col, i)).value = '%s' % (headers[x-1])
    ii = 2
    for record in records:  
        for x in range(1,len(record)+1):  
            col = get_column_letter(x)  
            ws.cell('%s%s'%(col, ii)).value = '%s' % (record[x-1])        
        ii+=1     
    ew.save(filename = fileName) 

def sql2lst(user, host, password, db, sqlSentence):
    import pymysql
    cnx = pymysql.connect(host = 'localhost',user = 'wjg',passwd = '0209',\
            db = 'managers',port = 3306,charset = 'utf8')
    cur = cnx.cursor()
    cur.execute(sqlSentence)
    records = []
    for r in cur:
        records.append(r)
    return records


def lst2sql(user, host, password, db, lst):
    import pymysql
    cnx = pymysql.connect(host = host,user = user,passwd = password,\
            db = db, port = 3306, charset = 'utf8')

    cur = cnx.cursor()
    tableName = input('输入表名：')
    create_table_sql = "CREATE TABLE if NOT EXISTS " + tableName + \
                        " (id int(10) AUTO_INCREMENT PRIMARY KEY,"
    print('共有' + str(len(lst)) + '列，' + str(len(lst[0])) + '行。')
    clmnName = []
    for i in range(len(lst)):
        clmnName.append(input('请输入第' + str(i+1) +  '列字段的字段名：'))
        create_table_sql = create_table_sql + " " +  clmnName[i] + ' '
        create_table_sql = create_table_sql + " " +  \
                           input('请输入第' + str(i+1) + \
                                 '列字段的mysql类型和宽度（如：int(5)、float(5,2)、varchar(10))：') + ','
    create_table_sql = create_table_sql[0:-1] +   ") CHARACTER SET utf8"
    cur.execute(create_table_sql)
    cnx.commit()
    insert_sql = "INSERT INTO " + tableName + " ("                                    
    for c in clmnName:
        insert_sql = insert_sql + c + ', '
    insert_sql = insert_sql[0:-2] + ") values "
    for i in range(len(lst[0])):
        a = "("
        for ii in range(len(lst)):  
            a = a + "'" + lst[ii][i] + "'" + ","
        sentence = insert_sql + a[:-1] + ')'
        print(sentence)
        cur.execute(sentence)
    cnx.commit()
    cnx.close()


def xl2sql(host, user, password, db, xl, clmnNumber, rowNumber, \
           tableNumber, sqlTablename):
    """xl2sql('localhost','wjg','0209','gsz','C:\\x.xlsm',316,6,1,'y')"""
    import pymysql
    lst = []
    import xlrd
    data = xlrd.open_workbook(xl)
    table = data.sheets()[tableNumber - 1]
    for i in range(clmnNumber):
        lst.append(table.col_values(i))
    cnx = pymysql.connect(host = host,user = user,passwd = password,\
            db = db, port = 3306, charset = 'utf8')
    cur = cnx.cursor()
    create_table_sql = "CREATE TABLE " + sqlTablename + \
                        " (id int(10) AUTO_INCREMENT PRIMARY KEY,"
    clmnName = []
    for i in range(len(lst)):
        clmnName.append(lst[i][rowNumber])
        create_table_sql = create_table_sql + " " +  clmnName[i] + ' '
        create_table_sql = create_table_sql + " " +  lst[i][rowNumber + 1] + ','
    create_table_sql = create_table_sql[0:-1] +   ") CHARACTER SET utf8"
    cur.execute(create_table_sql)
    cnx.commit()
    insert_sql = "INSERT INTO " + sqlTablename+ " ("                           
    for c in clmnName:
        insert_sql = insert_sql + c + ', '
    insert_sql = insert_sql[0:-2] + ") values "
    for i in range(2 + rowNumber, len(lst[0])):
        a = "("
        for ii in range(len(lst)):  
            a = a + "'" + str(lst[ii][i]) + "'" + ","
        sentence = insert_sql + a[:-1] + ')'
        cur.execute(sentence)
        cnx.commit()
    cnx.close()
    
def transpose(A):
    B = []
    for i in range(len(A[0])):
        t = []
        for ii in range(len(A)):
            t.append(A[ii][i])
        B.append(t)
    return B

def ftpUpload(fold1,address,name,code,fold2):
    '''ftpUpload('C:\\JR\\投资\\web\\上传\\fxbg\\','hz163423.ftp.aliapp.com','hz163423','Sien730123','/htdocs/fxbg/')'''
    from ftplib import FTP
    print("开始上传文件。")
    ftp=FTP()
    ftp.set_debuglevel(0)
    ftp.connect(address,21)
    ftp.login(name,code)
    bufsize = 1024
    ftp.cwd(fold2)
    files = dir2lst(fold1)
    for i in range(len(files)):
        file_handler = open(files[i],'rb')
        ftp.storbinary('STOR %s' % os.path.basename(files[i]),file_handler,bufsize)
        print(i) 
    file_handler.close()
    ftp.quit()
    
    
    
def pe(g, r, y1, y2, d, gdp):
    '''计算增长率对应的PE'''
    if g < -0.48:
        pe = 1
    else:
        for x in range(1,2500):
            pe1 = x / 10
            if g != r and r != d * g and r != gdp:
            
                A = (((1 / pe1) * (1 + g)) / (r - g)) * (1 - ((1 + g) / (1 + r)) ** y1)
                B = ((((1 / pe1) * ((1 + g) ** y1) * (1 + d * g)) / (r - d * g)) * ((1 - (1 + d * g) / (1 + r)) ** y2)) / ((1 + r) ** y1)
                C = (((1 / pe1) * ((1 + g) ** y1) * (1 + d * g) ** y2) / (r - gdp)) / ((1 + r) ** (y1 + y2))
                ABC = A + B + C
                if abs(ABC - 1) < 0.1:
                    break
        if pe1 < 5:
            pe = pe1
        else:
            pe = pe1
    return pe



def zzl(x1, x2):
    '''计算增长率，从0到正为20%，从0到负是-20%，从小于零到0，是20%，最大不超过200%，最小不小于-80%'''
    if x1 == 0 and x2 > 0 :
        r = 0.2
    elif x1 == 0 and x2 < 0 :        
        r = -0.2        
    elif x1 == 0 and x2 == 0 :        
        r = 0        
    elif x1 < 0 and x2 == 0 :     
        r = 0.2       
    else:
        r = (x2 - x1) / x1   
    if r > 2 :
        r = 2   
    elif r < -0.8 :
        r = -0.8
    return r
    
    
def tnn(t):
    '''文本数据转化为整数格式文本'''
    r = str(round(float(t)))
    return r
def tn(t):
    '''文本数据转化为保留一位小数格式文本'''
    r = str(round(float(t),1))
    return r
def tn2(t):
    '''文本数据转化为保留两位小数格式文本'''
    r = str(round(float(t),2))
    return r
def tp(t):
    '''文本数据转化为保留一位小数百分数格式文本'''
    r = str(round(float(t)*100,1)) + '%'
    return r
def tp2(t):
    '''文本数据转化为保留两位小数百分数格式文本'''
    r = str(round(float(t)*100,2)) + '%'
    return r


def sql(sql):
    import pymysql
    host = 'rds48z5k409g0753v62j.mysql.rds.aliyuncs.com'
    user = 'r0gbqnbc4x'
    password = '19760209'
    db = 'r0gbqnbc4x'
    cnx = pymysql.connect(host = host, user = user, passwd = password,\
            db = db, port = 3306, charset = 'utf8')
    cur = cnx.cursor()
    cur.execute(sql)
    for x in cur:
        print(x)
    return cur